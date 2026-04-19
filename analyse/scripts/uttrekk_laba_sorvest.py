"""
Stratifisert uttrekk av L-aba-hendelser (ABA løst av 110) for 110 Sør-Vest 2025,
for manuell dybdeanalyse i LEO-loggen.

- 50 hendelser totalt
- 4 per måned (10 mnd) + 5 per måned (2 mnd) = 50
- Tilfeldig utvalg innen hver måned (fast seed for reproduserbarhet)

Output: analyse/uttrekk/laba_sorvest_2025_dybdeanalyse.xlsx
"""
import pathlib
import random
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT = pathlib.Path(__file__).resolve().parent.parent.parent
DSB_FIL = PROJECT / "004 data" / "2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx"
OUT_DIR = PROJECT / "analyse" / "uttrekk"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FIL = OUT_DIR / "laba_sorvest_2025_dybdeanalyse.xlsx"

N_TOTAL = 50
SEED = 20260418  # dagens dato — reproduserbar

# === 1. LAST DATA ===
print("Laster DSB 2025-data...")
df = pd.read_excel(DSB_FIL, engine="openpyxl")
print(f"  {len(df):,} rader totalt")

# === 2. FILTRER TIL SØR-VEST ===
df["sentral_norm"] = df["110-sentral"].astype(str).str.strip()
sv = df[df["sentral_norm"].str.contains("Sør-Vest", na=False)].copy()
print(f"  Sør-Vest: {len(sv):,} rader")

# === 3. KATEGORISER (V3 — samme logikk som konflikt_total_belastning.py) ===
# Konverter Ressurs varslet til datetime
sv["Ressurs varslet"] = pd.to_datetime(sv["Ressurs varslet"], errors="coerce")

def klassifiser(row):
    """V3-regel (oppdatert 2026-04-19): L-aba krever Kilde=Alarm.
    ABA-oppdrag med Kilde=Samtale eller blank reklassifiseres til L-hendelse.
    MERK: Eksisterende utfylt datasett (50 hendelser, fra forrige utvalg) skal
    IKKE re-samples — denne oppdateringen er for konsistens med andre skript."""
    if pd.notna(row["Ressurs varslet"]):
        return "D"
    ot = str(row["Oppdragstype"]).strip() if pd.notna(row["Oppdragstype"]) else ""
    oot = str(row["Opprinnelig oppdragstype"]).strip() if pd.notna(row["Opprinnelig oppdragstype"]) else ""
    kilde = str(row["Kilde"]).strip() if "Kilde" in row.index and pd.notna(row["Kilde"]) else ""
    if ot == "Service":
        return "S"
    if ot in ("Nødanrop feilring", "Ikke reell nødmelding",
              "ECall feil bruk", "ECall teknisk/ukjent", "ECall veihjelp") or "feilring" in ot.lower():
        return "F"
    if "viderevarslet" in ot.lower() or "viderekoble" in ot.lower():
        return "V"
    if "ppdrag" in ot and "110" in ot:
        if oot == "ABA" and kilde == "Alarm":
            return "L-aba"
        if oot == "ABA":
            return "L-hendelse"
        if oot and oot.lower() != "nan":
            return "L-hendelse"
        return "L-ukjent"
    return "L-ukjent"

sv["v3_kat"] = sv.apply(klassifiser, axis=1)
laba = sv[sv["v3_kat"] == "L-aba"].copy()
print(f"  L-aba for Sør-Vest: {len(laba):,} rader")

# === 4. KLARGJØR FELTER ===
laba["Dato anrop"] = pd.to_datetime(laba["Dato anrop"], format="%d.%m.%Y", errors="coerce")
# Fallback hvis allerede datetime
if laba["Dato anrop"].isna().all():
    laba["Dato anrop"] = pd.to_datetime(laba["Dato anrop"], errors="coerce")
laba["maaned"] = laba["Dato anrop"].dt.month
laba["ukedag_navn"] = laba["Dato anrop"].dt.day_name(locale="nb_NO") if False else laba["Dato anrop"].dt.day_name()
laba["helg"] = laba["Dato anrop"].dt.weekday.isin([5, 6]).map({True: "Helg", False: "Hverdag"})

# Norske ukedager
UKEDAGER_NO = {
    "Monday": "Mandag", "Tuesday": "Tirsdag", "Wednesday": "Onsdag",
    "Thursday": "Torsdag", "Friday": "Fredag", "Saturday": "Lørdag", "Sunday": "Søndag",
}
laba["ukedag_no"] = laba["ukedag_navn"].map(UKEDAGER_NO)

# === 5. STRATIFISERT UTVALG ===
print("\nFordeling L-aba per måned i populasjonen:")
for m in range(1, 13):
    n = (laba["maaned"] == m).sum()
    print(f"  Måned {m:2d}: {n}")

# 10 mnd × 4 + 2 mnd × 5 = 50
random.seed(SEED)
np.random.seed(SEED)

maaneder = list(range(1, 13))
# Velg 2 tilfeldige måneder som får 5 hendelser (resten får 4)
ekstra_mnd = sorted(random.sample(maaneder, 2))
print(f"\nMåneder som får 5 hendelser: {ekstra_mnd}")

utvalg = []
for m in maaneder:
    n = 5 if m in ekstra_mnd else 4
    pool = laba[laba["maaned"] == m]
    if len(pool) < n:
        print(f"ADVARSEL måned {m}: kun {len(pool)} L-aba tilgjengelig, tar alle")
        utvalg.append(pool)
    else:
        utvalg.append(pool.sample(n=n, random_state=SEED + m))

utvalg_df = pd.concat(utvalg, ignore_index=True)
utvalg_df = utvalg_df.sort_values(["Dato anrop", "Time på døgnet"]).reset_index(drop=True)
print(f"\nTotalt uttrekk: {len(utvalg_df)} hendelser")

# === 6. BYGG OUTPUT ===
OUTPUT_KOLONNER = [
    "Nr",
    "Oppdrag ID",
    "110 ID",
    "Dato",
    "Tid anrop (ref)",
    "Oppdrag opprettet (ref)",
    "Ukedag",
    "Helg/Hverdag",
    "Kilde",
    "Opprinnelig type",
    "T_alarm_inn",
    "T_nødtelefon_inn",
    "T_avklart",
    "T_operatør_frigjort",
    "Bindingstid (min)",
    "Nødtelefon mottatt? (J/N)",
    "Kommentar",
]

# Konverter Oppdrag opprettet til datetime og format
oo = pd.to_datetime(utvalg_df["Oppdrag opprettet"], errors="coerce")
utvalg_df["oppdrag_opprettet_fmt"] = oo.dt.strftime("%H:%M:%S").fillna("")

# Tid anrop — kan være string "hh:mm:ss" eller time
def fmt_tid_anrop(v):
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return ""
    return s

out_df = pd.DataFrame({
    "Nr": range(1, len(utvalg_df) + 1),
    "Oppdrag ID": utvalg_df["Oppdrag ID"].astype(str),
    "110 ID": utvalg_df["110 ID"].astype(str),
    "Dato": utvalg_df["Dato anrop"].dt.strftime("%d.%m.%Y"),
    "Tid anrop (ref)": utvalg_df["Tid anrop"].apply(fmt_tid_anrop),
    "Oppdrag opprettet (ref)": utvalg_df["oppdrag_opprettet_fmt"],
    "Ukedag": utvalg_df["ukedag_no"],
    "Helg/Hverdag": utvalg_df["helg"],
    "Kilde": utvalg_df["Kilde"],
    "Opprinnelig type": utvalg_df["Opprinnelig oppdragstype"],
    "T_alarm_inn": "",
    "T_nødtelefon_inn": "",
    "T_avklart": "",
    "T_operatør_frigjort": "",
    "Bindingstid (min)": "",
    "Nødtelefon mottatt? (J/N)": "",
    "Kommentar": "",
})

# === 7. SKRIV XLSX MED FORMATERING ===
print(f"\nSkriver {OUT_FIL}")

wb = Workbook()
ws = wb.active
ws.title = "L-aba dybdeanalyse"

# Tittel-rad (merge over alle kolonner)
ws.cell(row=1, column=1).value = "L-aba dybdeanalyse — 110 Sør-Vest 2025"
ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="1A3A5C")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_KOLONNER))

ws.cell(row=2, column=1).value = (
    "Utvalg på 50 hendelser, stratifisert 4 per måned (10 mnd) + 5 per måned (2 mnd). "
    "Fyll ut tidskolonnene manuelt fra LEO-loggen. Bindingstid beregnes automatisk fra T_alarm_inn → T_operatør_frigjort."
)
ws.cell(row=2, column=1).font = Font(size=10, italic=True, color="555555")
ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(OUTPUT_KOLONNER))
ws.row_dimensions[2].height = 32

# Header (rad 4)
HEADER_ROW = 4
for ci, kol in enumerate(OUTPUT_KOLONNER, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=kol)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor="2A5A80")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

ws.row_dimensions[HEADER_ROW].height = 38

# Data-rader
DATA_START = HEADER_ROW + 1
for ri, row in enumerate(out_df.itertuples(index=False), start=DATA_START):
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=ri, column=ci, value=val if val != "" else None)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="center", wrap_text=False)
        # Zebra-rader (annenhver)
        if (ri - DATA_START) % 2 == 1:
            c.fill = PatternFill("solid", fgColor="F5F7FA")
        c.border = Border(
            left=Side(style="hair", color="DDDDDD"),
            right=Side(style="hair", color="DDDDDD"),
            top=Side(style="hair", color="DDDDDD"),
            bottom=Side(style="hair", color="DDDDDD"),
        )

    # Formel for bindingstid — kolonne 15 (O) = Bindingstid (min)
    # T_alarm_inn = kol 11 (K), T_operatør_frigjort = kol 14 (N)
    alarm_col = get_column_letter(11)  # K
    frigjort_col = get_column_letter(14)  # N
    ws.cell(row=ri, column=15).value = (
        f'=IF(AND(ISNUMBER({alarm_col}{ri}),ISNUMBER({frigjort_col}{ri})),'
        f'({frigjort_col}{ri}-{alarm_col}{ri})*1440,"")'
    )
    ws.cell(row=ri, column=15).number_format = "0.00"

# Kolonnebredder
BREDDER = {
    "Nr": 5, "Oppdrag ID": 12, "110 ID": 16, "Dato": 11,
    "Tid anrop (ref)": 14, "Oppdrag opprettet (ref)": 20,
    "Ukedag": 10, "Helg/Hverdag": 12, "Kilde": 10, "Opprinnelig type": 14,
    "T_alarm_inn": 13, "T_nødtelefon_inn": 16, "T_avklart": 13,
    "T_operatør_frigjort": 16, "Bindingstid (min)": 16,
    "Nødtelefon mottatt? (J/N)": 20, "Kommentar": 40,
}
for ci, kol in enumerate(OUTPUT_KOLONNER, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = BREDDER.get(kol, 12)

# Frys toppen
ws.freeze_panes = ws.cell(row=DATA_START, column=4)

# Gi manuell-utfylling-tidskolonnene tid-format (K, L, M, N)
for col_letter in ("K", "L", "M", "N"):
    for ri in range(DATA_START, DATA_START + len(out_df)):
        ws[f"{col_letter}{ri}"].number_format = "HH:MM:SS"

# Sheet 2: Oppsummering
ws2 = wb.create_sheet("Oppsummering")
ws2["A1"] = "Oppsummering L-aba dybdeanalyse — 110 Sør-Vest 2025"
ws2["A1"].font = Font(size=13, bold=True, color="1A3A5C")
ws2.merge_cells("A1:C1")

ws2["A3"] = "Populasjon (L-aba Sør-Vest 2025):"
ws2["A3"].font = Font(bold=True)
ws2["B3"] = len(laba)

ws2["A4"] = "Utvalgsstørrelse:"
ws2["A4"].font = Font(bold=True)
ws2["B4"] = len(out_df)

ws2["A5"] = "Utvalgsandel:"
ws2["A5"].font = Font(bold=True)
ws2["B5"] = f"{len(out_df)/len(laba)*100:.2f}%"

ws2["A7"] = "Fordeling per måned:"
ws2["A7"].font = Font(bold=True)
ws2["A8"] = "Måned"
ws2["B8"] = "Populasjon"
ws2["C8"] = "Utvalg"
for c in ("A8", "B8", "C8"):
    ws2[c].font = Font(bold=True, color="FFFFFF")
    ws2[c].fill = PatternFill("solid", fgColor="2A5A80")

MND_NAVN = ["Jan", "Feb", "Mar", "Apr", "Mai", "Jun",
            "Jul", "Aug", "Sep", "Okt", "Nov", "Des"]
for m in range(1, 13):
    ws2.cell(row=8 + m, column=1, value=MND_NAVN[m-1])
    ws2.cell(row=8 + m, column=2, value=int((laba["maaned"] == m).sum()))
    ws2.cell(row=8 + m, column=3, value=int((utvalg_df["maaned"] == m).sum()))

ws2["A22"] = "Reproduserbarhet:"
ws2["A22"].font = Font(bold=True)
ws2["A23"] = f"Random seed: {SEED}"
ws2["A24"] = "Måneder med 5 hendelser: " + ", ".join(MND_NAVN[m-1] for m in ekstra_mnd)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 15

# Sheet 3: Instruksjon
ws3 = wb.create_sheet("Instruksjon")
ws3["A1"] = "Instruksjon for manuell utfylling"
ws3["A1"].font = Font(size=13, bold=True, color="1A3A5C")

tekst = [
    "",
    "Formål:",
    "Kvantifisere faktisk operatørbindingstid for L-aba-hendelser (ABA løst av 110 uten utrykning) "
    "ved 110 Sør-Vest. Resultatet brukes til å validere eller justere bindingstidsestimatet på 3 min "
    "som ligger i kapasitetsmodellen.",
    "",
    "For hver hendelse i hoved-arket:",
    "",
    "1. Åpne hendelsen i LEO (bruk 110 ID for å finne den raskt).",
    "2. Fyll inn tidspunkter som klokkeslett (hh:mm:ss):",
    "   • T_alarm_inn — når ABA-signalet mottas i LEO",
    "   • T_nødtelefon_inn — når eventuell nødtelefon fra stedet besvares (90-sek-regel)",
    "   • T_avklart — når operatøren bekrefter ufarlig årsak (matlaging, damp, service mm.)",
    "   • T_operatør_frigjort — når oppdraget lukkes og operatør er tilgjengelig for neste",
    "3. Merk J/N om nødtelefon faktisk kom innen 90 sek.",
    "4. Bindingstid beregnes automatisk fra T_alarm_inn til T_operatør_frigjort.",
    "5. Bruk kommentarfeltet til observasjoner som kan forklare avvik (f.eks. samtidige "
    "hendelser, innringer som nekter, tvil om kategorisering).",
    "",
    "Merk:",
    "• Hvis hendelsen ikke matcher L-aba-definisjonen (ABA som lukkes uten utrykning), "
    "noter det i kommentar — vi vil se på klassifiseringsfeil.",
    "• Presisjon på sekund-nivå er ikke nødvendig; minutt-nivå er godt nok for analysen.",
    "• Returner arket ferdig utfylt til rune.grodemm@himolde.no.",
]
for i, t in enumerate(tekst, start=2):
    c = ws3.cell(row=i, column=1, value=t)
    c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[i].height = 20 if t and len(t) > 80 else 15

ws3.column_dimensions["A"].width = 110

wb.save(OUT_FIL)
print(f"Ferdig: {OUT_FIL}")

# === 8. KONSOLL-OPPSUMMERING ===
print("\nUtvalg per måned:")
for m in range(1, 13):
    n = (utvalg_df["maaned"] == m).sum()
    print(f"  Måned {m:2d}: {n}")

print(f"\nUkedagsfordeling:")
for d, n in utvalg_df["ukedag_no"].value_counts().items():
    print(f"  {d}: {n}")

print(f"\nHelg/hverdag:")
for d, n in utvalg_df["helg"].value_counts().items():
    print(f"  {d}: {n}")
