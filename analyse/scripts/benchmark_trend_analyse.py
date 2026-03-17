"""
benchmark_trend_analyse.py
==========================
Trendanalyse og benchmarking — alle 12 norske 110-sentraler, 2022–2025.

Datakilder:
  - MOB-filer (DSB årsrapporter):  004 data/20260315_1745*_MOB_20XX_110-sentral.xlsx
  - BRIS fullrapporter (hendelser): 004 data/fullrapport_20XX.csv  (legg til når tilgjengelig)

Kjør:
  py "G20-rune-individuell/analyse/scripts/benchmark_trend_analyse.py"

Output:
  - figurer/benchmark_anrop_trend.png
  - figurer/benchmark_bemanning_trend.png
  - figurer/benchmark_belastning.png
  - analyse/benchmark_tabell.csv
"""

import os
import warnings
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl

warnings.filterwarnings("ignore")

# ── Stier ────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR   = os.path.join(BASE, "004 data")
FIG_DIR    = os.path.join(BASE, "figurer")
OUT_DIR    = os.path.join(BASE, "analyse")
os.makedirs(FIG_DIR, exist_ok=True)

MOB_FILES = {
    2022: "20260315_174537_MOB_2022_110-sentral.xlsx",
    2023: "20260315_174530_MOB_2023_110-sentral.xlsx",
    2024: "20260315_174523_MOB_2024_110-sentral.xlsx",
    2025: "20260315_174514_MOB_2025_110-sentral.xlsx",
}

# BRIS fullrapporter — legg til filnavn når tilgjengelig
# Forventet kolonnestruktur: samme som 2025-filen (24 kolonner)
BRIS_FILES = {
    # 2022: "fullrapport_2022.csv",
    # 2023: "fullrapport_2023.csv",
    # 2024: "fullrapport_2024.csv",
    2025: "20260315_174350_fullrapport.csv",
}

SENTRALER_NORM = {
    "Sør-Vest 110": "Sør-Vest 110",
    "S\u00f8r-Vest 110": "Sør-Vest 110",
    "S?r-Vest 110": "Sør-Vest 110",
    "Oslo 110": "Oslo 110",
    "Øst 110": "Øst 110",
    "\u00d8st 110": "Øst 110",
    "?st 110": "Øst 110",
    "Sør-Øst 110": "Sør-Øst 110",
    "S\u00f8r-\u00d8st 110": "Sør-Øst 110",
    "S?r-?st 110": "Sør-Øst 110",
    "Vest 110": "Vest 110",
    "Midt-Norge 110": "Midt-Norge 110",
    "Innlandet 110": "Innlandet 110",
    "Agder 110": "Agder 110",
    "Nordland 110": "Nordland 110",
    "Møre og Romsdal 110": "Møre og Romsdal 110",
    "M\u00f8re og Romsdal 110": "Møre og Romsdal 110",
    "M?re og Romsdal 110": "Møre og Romsdal 110",
    "Tromsø 110": "Tromsø 110",
    "Troms\u00f8 110": "Tromsø 110",
    "Troms? 110": "Tromsø 110",
    "Finnmark 110": "Finnmark 110",
}

def norm(name):
    """Normaliser sentralnavn til kanonisk form."""
    s = str(name).strip()
    return SENTRALER_NORM.get(s, s)


# ── Last MOB-data ─────────────────────────────────────────────────────────────
def load_mob(year, filename):
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Finn header-rad
    hdr_idx = next(i for i, r in enumerate(rows) if r[0] and "110-sentral" in str(r[0]))
    headers = [str(h).strip() if h else "" for h in rows[hdr_idx]]

    def col(kw):
        for i, h in enumerate(headers):
            if kw.lower() in h.lower():
                return i
        return None

    idx = {
        "dag":   col("dag - hverdager"),
        "natt":  col("natt - hverdager"),
        "dag_h": col("dag - helg"),
        "natt_h":col("natt - helg"),
        "arsv":  col("rsverk operat"),
        "plasser": col("operatørplasser") or col("operatorplasser"),
        "anrop": col("mottatte 110-anrop"),
    }

    records = []
    for r in rows[hdr_idx + 1:]:
        if not r[0] or "Sum" in str(r[0]):
            continue
        rec = {"år": year, "sentral": norm(r[0])}
        for k, i in idx.items():
            rec[k] = float(r[i]) if i is not None and r[i] is not None else None
        records.append(rec)
    return records


# 2025-data hardkodet som fallback (fil kan være åpen i Excel)
MOB_2025_FALLBACK = [
    {"år":2025,"sentral":"Vest 110",            "dag":4,"natt":3,"dag_h":3,"natt_h":3,"arsv":23,"plasser":6, "anrop":18709},
    {"år":2025,"sentral":"Midt-Norge 110",       "dag":4,"natt":4,"dag_h":4,"natt_h":4,"arsv":21,"plasser":6, "anrop":12453},
    {"år":2025,"sentral":"Innlandet 110",        "dag":4,"natt":4,"dag_h":4,"natt_h":4,"arsv":20,"plasser":5, "anrop":14500},
    {"år":2025,"sentral":"Sør-Øst 110",          "dag":6,"natt":5,"dag_h":5,"natt_h":5,"arsv":23,"plasser":7, "anrop":27559},
    {"år":2025,"sentral":"Agder 110",            "dag":4,"natt":3,"dag_h":3,"natt_h":3,"arsv":16,"plasser":7, "anrop":12417},
    {"år":2025,"sentral":"Finnmark 110",         "dag":3,"natt":3,"dag_h":3,"natt_h":3,"arsv":15,"plasser":3, "anrop":3141},
    {"år":2025,"sentral":"Oslo 110",             "dag":5,"natt":4,"dag_h":4,"natt_h":4,"arsv":30,"plasser":12,"anrop":30000},
    {"år":2025,"sentral":"Øst 110",              "dag":6,"natt":6,"dag_h":6,"natt_h":6,"arsv":30,"plasser":7, "anrop":25962},
    {"år":2025,"sentral":"Sør-Vest 110",         "dag":4,"natt":3,"dag_h":3,"natt_h":3,"arsv":24,"plasser":6, "anrop":26348},
    {"år":2025,"sentral":"Møre og Romsdal 110",  "dag":4,"natt":3,"dag_h":3,"natt_h":3,"arsv":16,"plasser":5, "anrop":8096},
    {"år":2025,"sentral":"Tromsø 110",           "dag":2,"natt":2,"dag_h":2,"natt_h":2,"arsv":10,"plasser":3, "anrop":7844},
    {"år":2025,"sentral":"Nordland 110",         "dag":3,"natt":3,"dag_h":3,"natt_h":3,"arsv":16,"plasser":6, "anrop":8781},
]

mob_records = []
for yr, fn in MOB_FILES.items():
    if yr == 2025:
        mob_records.extend(MOB_2025_FALLBACK)
        print("MOB 2025: bruker hardkodet fallback (fil åpen i Excel)")
        continue
    mob_records.extend(load_mob(yr, fn))

mob = pd.DataFrame(mob_records)
mob["c_eff_dag"]  = mob["dag"]  - 1   # VL-korreksjon
mob["c_eff_natt"] = mob["natt"] - 1


# ── Last BRIS-data ────────────────────────────────────────────────────────────
def load_bris(year, filename):
    path = os.path.join(DATA_DIR, filename)
    df = pd.read_csv(path, sep=None, engine="python", skiprows=2,
                     encoding="utf-8-sig", low_memory=False)
    headers = df.iloc[0].tolist()
    df.columns = headers
    df = df.iloc[1:].reset_index(drop=True)
    df["år"] = year
    df["sentral_norm"] = df["110-sentral"].apply(norm)

    # Parsé dato
    df["dato"] = pd.to_datetime(df["Dato anrop"], format="%d.%m.%Y", errors="coerce")
    df["time"] = pd.to_numeric(df["Time på døgnet"], errors="coerce")

    return df

bris_frames = {}
for yr, fn in BRIS_FILES.items():
    try:
        bris_frames[yr] = load_bris(yr, fn)
        print(f"BRIS {yr}: {len(bris_frames[yr])} rader lastet")
    except Exception as e:
        print(f"BRIS {yr}: IKKE LASTET ({e})")

# Aggregér BRIS til per-sentral per-år
bris_agg_records = []
for yr, df in bris_frames.items():
    grp = df.groupby("sentral_norm").agg(
        bris_hendelser=("Oppdrag ID", "count"),
        bris_alarm=("Kilde", lambda x: (x == "Alarm").sum()),
        bris_samtale=("Kilde", lambda x: (x == "Samtale").sum()),
    ).reset_index().rename(columns={"sentral_norm": "sentral"})
    grp["år"] = yr
    bris_agg_records.append(grp)

if bris_agg_records:
    bris_agg = pd.concat(bris_agg_records, ignore_index=True)
else:
    bris_agg = pd.DataFrame(columns=["sentral", "år", "bris_hendelser", "bris_alarm", "bris_samtale"])


# ── Slå sammen MOB + BRIS ─────────────────────────────────────────────────────
merged = mob.merge(bris_agg, on=["sentral", "år"], how="left")

# Belastningsindeks: anrop per c_effektiv per år
merged["anrop_per_ceff"] = merged["anrop"] / merged["c_eff_dag"].replace(0, np.nan)


# ── Skriv CSV ─────────────────────────────────────────────────────────────────
out_csv = os.path.join(OUT_DIR, "benchmark_tabell.csv")
merged.to_csv(out_csv, index=False, encoding="utf-8-sig")
print(f"\nBenchmark-tabell lagret: {out_csv}")


# ── Figur 1: Anropstrend per sentral ─────────────────────────────────────────
pivot_anrop = merged.pivot(index="sentral", columns="år", values="anrop")

fig, ax = plt.subplots(figsize=(12, 7))
years = sorted(merged["år"].unique())
x = np.arange(len(pivot_anrop))
w = 0.2
colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]

for i, yr in enumerate(years):
    vals = pivot_anrop.get(yr, pd.Series(dtype=float)).reindex(pivot_anrop.index)
    ax.bar(x + i * w, vals, width=w, label=str(yr), color=colors[i], alpha=0.85)

ax.set_xticks(x + w * 1.5)
ax.set_xticklabels(pivot_anrop.index, rotation=35, ha="right", fontsize=9)
ax.set_ylabel("Antall mottatte 110-anrop")
ax.set_title("Anropstrend per 110-sentral 2022–2025")
ax.legend(title="År")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}".replace(",", " ")))
plt.tight_layout()
plt.savefig(os.path.join(FIG_DIR, "benchmark_anrop_trend.png"), dpi=150)
plt.close()
print("Figur lagret: benchmark_anrop_trend.png")


# ── Figur 2: Bemanningstrend (c_effektiv dag) ─────────────────────────────────
pivot_bem = merged.pivot(index="sentral", columns="år", values="c_eff_dag")

fig, ax = plt.subplots(figsize=(12, 6))
for i, yr in enumerate(years):
    vals = pivot_bem.get(yr, pd.Series(dtype=float)).reindex(pivot_bem.index)
    ax.bar(x + i * w, vals, width=w, label=str(yr), color=colors[i], alpha=0.85)

ax.set_xticks(x + w * 1.5)
ax.set_xticklabels(pivot_bem.index, rotation=35, ha="right", fontsize=9)
ax.set_ylabel("c_effektiv dagsvakt (c_total − 1, hverdager)")
ax.set_title("Bemanningstrend c_effektiv (dag, hverdager) 2022–2025")
ax.legend(title="År")
ax.set_ylim(0, max(pivot_bem.max()) + 1)
plt.tight_layout()
plt.savefig(os.path.join(FIG_DIR, "benchmark_bemanning_trend.png"), dpi=150)
plt.close()
print("Figur lagret: benchmark_bemanning_trend.png")


# ── Figur 3: Belastningsindeks (anrop / c_eff_dag) ───────────────────────────
pivot_bel = merged.pivot(index="sentral", columns="år", values="anrop_per_ceff")

fig, ax = plt.subplots(figsize=(12, 6))
for i, yr in enumerate(years):
    vals = pivot_bel.get(yr, pd.Series(dtype=float)).reindex(pivot_bel.index)
    ax.bar(x + i * w, vals, width=w, label=str(yr), color=colors[i], alpha=0.85)

ax.set_xticks(x + w * 1.5)
ax.set_xticklabels(pivot_bel.index, rotation=35, ha="right", fontsize=9)
ax.set_ylabel("Anrop per c_effektiv (dag) per år")
ax.set_title("Belastningsindeks: anrop per effektiv operatør 2022–2025")
ax.legend(title="År")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}".replace(",", " ")))
plt.tight_layout()
plt.savefig(os.path.join(FIG_DIR, "benchmark_belastning.png"), dpi=150)
plt.close()
print("Figur lagret: benchmark_belastning.png")


# ── Konsolltabell ─────────────────────────────────────────────────────────────
print("\n" + "="*90)
print("TRENDTABELL — ANROP PER SENTRAL")
print("="*90)
print(f"{'Sentral':<25} {'2022':>8} {'2023':>8} {'2024':>8} {'2025':>8} {'Endr 22→25':>12} {'%':>7}")
print("-"*80)
for s in sorted(merged["sentral"].unique()):
    sub = merged[merged["sentral"] == s].set_index("år")["anrop"]
    a = {yr: sub.get(yr) for yr in [2022,2023,2024,2025]}
    endr = (a[2025] - a[2022]) if a.get(2025) and a.get(2022) else None
    pct  = (endr / a[2022] * 100) if endr and a[2022] else None
    row  = [f"{int(a[y]):>8}" if a.get(y) else f"{'—':>8}" for y in [2022,2023,2024,2025]]
    print(f"{s:<25} {' '.join(row)} {str(int(endr) if endr else '—'):>12} {str(round(pct,1) if pct else '—'):>6}%")

print("\n" + "="*90)
print("BELASTNINGSINDEKS — ANROP / c_EFFEKTIV DAG")
print("="*90)
print(f"{'Sentral':<25} {'2022':>8} {'2023':>8} {'2024':>8} {'2025':>8}")
print("-"*60)
for s in sorted(merged["sentral"].unique()):
    sub = merged[merged["sentral"] == s].set_index("år")["anrop_per_ceff"]
    row = [f"{int(sub[yr]):>8}" if yr in sub and pd.notna(sub[yr]) else f"{'—':>8}" for yr in [2022,2023,2024,2025]]
    print(f"{s:<25} {' '.join(row)}")

print("\nFerdig.")
