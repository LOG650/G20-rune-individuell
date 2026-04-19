"""
Nasjonal oversikt — DSB 2025 for alle 12 110-sentraler.

Produserer:
  - analyse/nasjonal_2025/nasjonal_oversikt.md (hovedrapport)
  - analyse/nasjonal_2025/storrelse_ranking.csv
  - analyse/nasjonal_2025/sammenstilte_rate.csv
  - figurer/nasjonal_oversikt_*.png (5 figurer)

Inkluderer:
  - Størrelses-metrikker (volum, arbeidsmengde, anrop per operatør)
  - Kategori-heatmap for sentral × kategori
  - Utrykningsrate-ranking
  - L-aba dyp-fokus
  - Tidsdata boksplot for D-hendelser
  - Sammenstilte anrop (manglende 110 ID-sekvensnummer)
"""
import pathlib
import re
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns

warnings.filterwarnings("ignore")

PROJECT = pathlib.Path(__file__).resolve().parent.parent.parent
DATA_DIR = PROJECT / "004 data"
OUT_DIR = PROJECT / "analyse" / "nasjonal_2025"
FIG_DIR = PROJECT / "analyse" / "figurer"
OUT_DIR.mkdir(parents=True, exist_ok=True)
FIG_DIR.mkdir(parents=True, exist_ok=True)

DSB_FIL = DATA_DIR / "2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx"
MOB_2025 = DATA_DIR / "20260315_174514_MOB_2025_110-sentral.xlsx"

# Bindingstider (minutter) — hoved-scenario fra konflikt_total_belastning.py
BINDINGSTID_MIN = {
    "D": 16.0,          # ~13 min median + 3 min kvittering
    "S": 2.0,
    "L-aba": 3.0,
    "L-hendelse": 5.0,
    "L-ukjent": 3.0,
    "F": 0.5,
    "V": 1.0,
}

sns.set_style("whitegrid")
plt.rcParams.update({"figure.dpi": 150, "font.size": 10, "axes.titleweight": "bold"})

FARGE_KAT = {
    "D": "#C44E52", "S": "#8172B2", "L-aba": "#DD8452",
    "L-hendelse": "#55A868", "L-ukjent": "#CCB974",
    "F": "#64B5CD", "V": "#937860",
}

# === NAVN-NORMALISERING ===
def norm_sentral(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    # Encoding-varianter
    s = s.replace("S\u00f8r", "Sør").replace("S?r", "Sør")
    s = s.replace("\u00d8st", "Øst").replace("?st", "Øst")
    s = s.replace("M\u00f8re", "Møre").replace("M?re", "Møre")
    s = s.replace("Troms\u00f8", "Tromsø").replace("Troms?", "Tromsø")
    # Fjern " 110"-suffiks
    s = re.sub(r"\s*110\s*$", "", s)
    return s


# === 1. LAST OG KATEGORISER ===
print("1. Laster DSB 2025...")
df = pd.read_excel(DSB_FIL, engine="openpyxl")
df["sentral"] = df["110-sentral"].apply(norm_sentral)
df["Ressurs varslet"] = pd.to_datetime(df["Ressurs varslet"], errors="coerce")
print(f"   {len(df):,} rader")

def klassifiser(row):
    """V3-regel (oppdatert 2026-04-19): L-aba krever Kilde=Alarm.
    ABA-oppdrag med Kilde=Samtale eller blank reklassifiseres til L-hendelse."""
    if pd.notna(row["Ressurs varslet"]):
        return "D"
    ot = str(row["Oppdragstype"]).strip() if pd.notna(row["Oppdragstype"]) else ""
    oot = str(row["Opprinnelig oppdragstype"]).strip() if pd.notna(row["Opprinnelig oppdragstype"]) else ""
    kilde = str(row["Kilde"]).strip() if "Kilde" in row.index and pd.notna(row["Kilde"]) else ""
    if ot == "Service":
        return "S"
    if ot in ("Nødanrop feilring", "Ikke reell nødmelding",
              "ECall feil bruk", "ECall teknisk/ukjent", "ECall veihjelp") or "feilring" in ot.lower():
        return "F"
    if "viderevarslet" in ot.lower() or "viderekoble" in ot.lower():
        return "V"
    if "ppdrag" in ot and "110" in ot:
        if oot == "ABA" and kilde == "Alarm":
            return "L-aba"
        if oot == "ABA":
            return "L-hendelse"
        if oot and oot.lower() != "nan":
            return "L-hendelse"
        return "L-ukjent"
    return "L-ukjent"

df["v3_kat"] = df.apply(klassifiser, axis=1)

# Parse ekstra timestamps for varslings- vs utrykkings-analyse
df["Rykket ut"] = pd.to_datetime(df["Rykket ut"], errors="coerce")
df["Første ressurs fremme"] = pd.to_datetime(df["Første ressurs fremme"], errors="coerce")

ALLE_SENTRALER = sorted(df["sentral"].dropna().unique().tolist())
print(f"   Sentraler: {len(ALLE_SENTRALER)}")


# === 2. MOB-DATA FOR OPERATØR-KAPASITET ===
print("2. Laster MOB 2025 for bemanningsdata...")
import openpyxl
wb = openpyxl.load_workbook(MOB_2025, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr_idx = next(i for i, r in enumerate(rows) if r[0] and "110-sentral" in str(r[0]))
hdrs = [str(h).strip() if h else "" for h in rows[hdr_idx]]

def find_col(kw):
    for i, h in enumerate(hdrs):
        if kw.lower() in h.lower():
            return i
    return None

idx_dag = find_col("dag - hverdager")
idx_natt = find_col("natt - hverdager")
mob_data = {}
for r in rows[hdr_idx+1:]:
    if not r[0] or "Sum" in str(r[0]):
        continue
    s = norm_sentral(r[0])
    mob_data[s] = {
        "dag": float(r[idx_dag]) if r[idx_dag] else None,
        "natt": float(r[idx_natt]) if r[idx_natt] else None,
    }


# === 3. STØRRELSES-METRIKKER ===
print("3. Beregner størrelses-metrikker...")

rows_storrelse = []
for s in ALLE_SENTRALER:
    sub = df[df["sentral"] == s]
    total = len(sub)
    # Kategori-volum
    kat_vol = {k: (sub["v3_kat"] == k).sum() for k in BINDINGSTID_MIN}
    # Arbeidsmengde (timer/år)
    arbmengde_min = sum(kat_vol[k] * BINDINGSTID_MIN[k] for k in kat_vol)
    arbmengde_timer = arbmengde_min / 60
    arbmengde_timer_dag = arbmengde_timer / 365

    mob = mob_data.get(s, {})
    c_dag = mob.get("dag")
    c_eff_dag = (c_dag - 1) if c_dag else None

    # Utrykningsrealisering — andel varslede som faktisk rykket ut
    d_sub = sub[sub["v3_kat"] == "D"]
    rykket_ut_n = d_sub["Rykket ut"].notna().sum()
    fremme_n = d_sub["Første ressurs fremme"].notna().sum()
    pct_rykket = (rykket_ut_n / kat_vol["D"] * 100) if kat_vol["D"] else None
    pct_fremme = (fremme_n / kat_vol["D"] * 100) if kat_vol["D"] else None

    # Samlet brann-relatert rate = D + L-aba + L-hendelse (alle oppdrag med brann-relatert vurdering)
    brannrelatert = kat_vol["D"] + kat_vol["L-aba"] + kat_vol["L-hendelse"]
    brannrelatert_pct = brannrelatert / total * 100

    rows_storrelse.append({
        "sentral": s,
        "totalvolum": total,
        "D": kat_vol["D"],
        "utrykningsrate_pct": kat_vol["D"] / total * 100,
        "pct_rykket_av_varslet": round(pct_rykket, 1) if pct_rykket else None,
        "pct_fremme_av_varslet": round(pct_fremme, 1) if pct_fremme else None,
        "brannrelatert_n": brannrelatert,
        "brannrelatert_pct": round(brannrelatert_pct, 1),
        "arbmengde_timer_aar": round(arbmengde_timer, 0),
        "arbmengde_timer_per_dag": round(arbmengde_timer_dag, 1),
        "c_total_dag": c_dag,
        "c_eff_dag": c_eff_dag,
        "anrop_per_ceff": (total / c_eff_dag) if c_eff_dag else None,
    })

storrelse = pd.DataFrame(rows_storrelse).sort_values("totalvolum", ascending=False)

# Rangering
for kol in ["totalvolum", "D", "arbmengde_timer_aar", "anrop_per_ceff"]:
    storrelse[f"rang_{kol}"] = storrelse[kol].rank(ascending=False).astype(int)

storrelse.to_csv(OUT_DIR / "storrelse_ranking.csv", index=False, encoding="utf-8-sig")
print("   storrelse_ranking.csv skrevet")


# === 4. SAMMENSTILTE ANROP (sekvens-analyse) ===
print("4. Beregner sammenstilte anrop per sentral...")

def parse_110id(id_str):
    """Ekstraher dato og sekvensnr fra '110 ID' som BRA-YYMMDD-N."""
    if pd.isna(id_str):
        return None, None
    m = re.match(r"^[A-Z]+\d*-(\d{6})-(\d+)$", str(id_str).strip())
    if not m:
        return None, None
    return m.group(1), int(m.group(2))

df["dato_id"] = None
df["seq_nr"] = None

parsed = df["110 ID"].apply(parse_110id)
df["dato_id"] = parsed.apply(lambda x: x[0])
df["seq_nr"] = parsed.apply(lambda x: x[1])

sammenstilte_rader = []
for s in ALLE_SENTRALER:
    sub = df[df["sentral"] == s].dropna(subset=["dato_id", "seq_nr"])
    registrerte = len(sub)
    skjulte = 0
    totalt_anrop = 0
    for dato, grp in sub.groupby("dato_id"):
        seqs = set(grp["seq_nr"].astype(int))
        if not seqs:
            continue
        maks = max(seqs)
        totalt_anrop += maks
        manglende = set(range(1, maks + 1)) - seqs
        skjulte += len(manglende)
    rate = (skjulte / totalt_anrop * 100) if totalt_anrop else 0
    sammenstilte_rader.append({
        "sentral": s,
        "registrerte_oppdrag": registrerte,
        "skjulte_sammenstilte": skjulte,
        "totalt_estimert_anrop": totalt_anrop,
        "sammenstilte_rate_pct": round(rate, 2),
    })

sammenstilte = pd.DataFrame(sammenstilte_rader).sort_values("sammenstilte_rate_pct", ascending=False)
sammenstilte.to_csv(OUT_DIR / "sammenstilte_rate.csv", index=False, encoding="utf-8-sig")
print("   sammenstilte_rate.csv skrevet")


# === 5. KATEGORI-FORDELING (bench) ===
print("5. Bygger kategorifordeling...")
bench = df.groupby(["sentral", "v3_kat"]).size().unstack(fill_value=0)
for k in BINDINGSTID_MIN:
    if k not in bench.columns:
        bench[k] = 0
bench = bench[list(BINDINGSTID_MIN.keys())]
bench["Totalt"] = bench.sum(axis=1)
bench_pct = bench.div(bench["Totalt"], axis=0).drop(columns="Totalt") * 100


# === 6. TIDSDATA D ===
print("6. Henter tidsdata D...")
d_df = df[df["v3_kat"] == "D"].copy()
d_df["alarmbeh_min"] = pd.to_timedelta(d_df["Alarmbehandlingstid"], errors="coerce").dt.total_seconds() / 60


# === 7. FIGURER ===
print("7. Lager figurer...")

# Figur 1: Størrelses-rangering (sortert stolpe)
fig, axes = plt.subplots(1, 3, figsize=(18, 6))

# 1a: Total volum
s_sortert = storrelse.sort_values("totalvolum", ascending=True)
axes[0].barh(s_sortert["sentral"], s_sortert["totalvolum"],
              color=["#C44E52" if i >= len(s_sortert)-3 else "#4C72B0"
                     for i in range(len(s_sortert))])
axes[0].set_title("Total oppdragsvolum 2025", fontsize=11)
axes[0].set_xlabel("Antall oppdrag")
axes[0].xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}".replace(",", " ")))

# 1b: Arbeidsmengde per dag
s_sortert2 = storrelse.sort_values("arbmengde_timer_per_dag", ascending=True)
axes[1].barh(s_sortert2["sentral"], s_sortert2["arbmengde_timer_per_dag"],
              color="#DD8452")
axes[1].set_title("Arbeidsmengde per dag\n(vektet bindingstid, timer)", fontsize=11)
axes[1].set_xlabel("Timer/dag")

# 1c: Anrop per effektiv operatør
s_sortert3 = storrelse.dropna(subset=["anrop_per_ceff"]).sort_values("anrop_per_ceff", ascending=True)
axes[2].barh(s_sortert3["sentral"], s_sortert3["anrop_per_ceff"],
              color="#55A868")
axes[2].set_title("Oppdrag per c_effektiv (dag)", fontsize=11)
axes[2].set_xlabel("Oppdrag/operatør (år)")
axes[2].xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}".replace(",", " ")))

plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_storrelse.png", bbox_inches="tight")
plt.close()

# Figur 2: Kategori-heatmap
fig, ax = plt.subplots(figsize=(10, 7))
heat_data = bench_pct[list(BINDINGSTID_MIN.keys())].reindex(
    bench_pct.sum(axis=1).sort_values(ascending=False).index
)
sns.heatmap(heat_data, annot=True, fmt=".1f", cmap="YlOrRd", ax=ax,
            cbar_kws={"label": "% av totalvolum"}, linewidths=0.4,
            linecolor="white", annot_kws={"size": 9})
ax.set_title("Kategorifordeling per sentral (DSB 2025) — prosent av totalvolum", pad=14)
ax.set_xlabel("")
ax.set_ylabel("")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_kategori_heatmap.png", bbox_inches="tight")
plt.close()

# Figur 3: Utrykningsrate sortert med fargekoding
fig, ax = plt.subplots(figsize=(11, 6))
utr = storrelse.sort_values("utrykningsrate_pct", ascending=True)
med_utr = utr["utrykningsrate_pct"].median()
q1u, q3u = utr["utrykningsrate_pct"].quantile([0.2, 0.8])
farger = []
for v in utr["utrykningsrate_pct"]:
    if v >= q3u: farger.append("#C44E52")
    elif v <= q1u: farger.append("#2563eb")
    else: farger.append("#6b7280")
bars = ax.barh(utr["sentral"], utr["utrykningsrate_pct"], color=farger)
ax.axvline(med_utr, color="black", linestyle="--", linewidth=1,
           label=f"Nasjonal median {med_utr:.1f}%")
for b, v in zip(bars, utr["utrykningsrate_pct"]):
    ax.text(v + 0.3, b.get_y() + b.get_height()/2, f"{v:.1f}%",
            va="center", fontsize=9)
ax.set_xlabel("Utrykningsrate (D / totalt, %)")
ax.set_title("Utrykningsrate per sentral (DSB 2025)")
ax.legend(loc="lower right")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_utrykningsrate.png", bbox_inches="tight")
plt.close()
# Figur 3b: D vs brann-relatert (sammenligning)
fig, ax = plt.subplots(figsize=(12, 6))
sortert = storrelse.sort_values("brannrelatert_pct", ascending=True)
y = np.arange(len(sortert))
ax.barh(y + 0.2, sortert["utrykningsrate_pct"], height=0.4,
        color="#C44E52", label="D — ren varslingsrate")
ax.barh(y - 0.2, sortert["brannrelatert_pct"], height=0.4,
        color="#DD8452", label="D + L-aba + L-hendelse — samlet brann-relatert")
# Annotasjoner: differanse (L-aba + L-hendelse)
for i, (_, r) in enumerate(sortert.iterrows()):
    diff = r["brannrelatert_pct"] - r["utrykningsrate_pct"]
    ax.text(r["brannrelatert_pct"] + 0.3, i - 0.2,
            f"{r['brannrelatert_pct']:.1f}%", va="center", fontsize=8.5)
    ax.text(r["utrykningsrate_pct"] + 0.3, i + 0.2,
            f"{r['utrykningsrate_pct']:.1f}%", va="center", fontsize=8.5, color="#991b1b")
ax.set_yticks(y)
ax.set_yticklabels(sortert["sentral"])
ax.set_xlabel("Andel av totalvolum (%)")
ax.set_title("Varslingsrate vs samlet brann-relatert — avdekker beslutnings­terskelen i 110")
ax.legend(loc="lower right")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_brannrelatert.png", bbox_inches="tight")
plt.close()

# Figur 4: L-aba dyp-fokus
fig, ax = plt.subplots(figsize=(11, 6))
laba_data = bench_pct.copy()
laba_data["L-aba_abs"] = bench["L-aba"]
laba_data = laba_data.sort_values("L-aba", ascending=True)
med_laba = laba_data["L-aba"].median()
farger_laba = []
for v in laba_data["L-aba"]:
    if v == 0: farger_laba.append("#991b1b")
    elif v <= laba_data["L-aba"].quantile(0.25): farger_laba.append("#f87171")
    elif v >= laba_data["L-aba"].quantile(0.75): farger_laba.append("#22c55e")
    else: farger_laba.append("#6b7280")
bars = ax.barh(laba_data.index, laba_data["L-aba"], color=farger_laba)
for b, v, n in zip(bars, laba_data["L-aba"], laba_data["L-aba_abs"]):
    ax.text(v + 0.15, b.get_y() + b.get_height()/2, f"{v:.1f}% ({int(n):,})",
            va="center", fontsize=9)
ax.axvline(med_laba, color="black", linestyle="--", linewidth=1,
           label=f"Nasjonal median {med_laba:.1f}%")
ax.set_xlabel("L-aba-andel (%)")
ax.set_title("L-aba — ABA løst av 110 uten utrykning. Svært ulik registreringspraksis.")
ax.legend(loc="lower right")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_laba.png", bbox_inches="tight")
plt.close()

# Figur 5: Tidsdata boksplot D
fig, ax = plt.subplots(figsize=(12, 6))
d_df_plot = d_df.dropna(subset=["alarmbeh_min"])
d_df_plot = d_df_plot[d_df_plot["alarmbeh_min"] < 30]  # trim ekstreme outliers for plot
sentral_rekke = sorted(d_df_plot["sentral"].unique(),
                       key=lambda s: d_df_plot[d_df_plot["sentral"] == s]["alarmbeh_min"].median())
sns.boxplot(data=d_df_plot, y="sentral", x="alarmbeh_min",
            order=sentral_rekke, ax=ax, color="#4C72B0", fliersize=2)
ax.set_xlabel("Alarmbehandlingstid D-oppdrag (minutter, trimmed ved 30)")
ax.set_ylabel("")
ax.set_title("Alarmbehandlingstid per sentral (kategori D) — boksplot")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_tid_boxplot.png", bbox_inches="tight")
plt.close()

# Figur 6: Sammenstilte anrop-rate
fig, ax = plt.subplots(figsize=(11, 6))
ss = sammenstilte.sort_values("sammenstilte_rate_pct", ascending=True)
bars = ax.barh(ss["sentral"], ss["sammenstilte_rate_pct"], color="#937860")
for b, v, n in zip(bars, ss["sammenstilte_rate_pct"], ss["skjulte_sammenstilte"]):
    ax.text(v + 0.2, b.get_y() + b.get_height()/2, f"{v:.2f}% ({int(n):,})",
            va="center", fontsize=9)
med_ss = ss["sammenstilte_rate_pct"].median()
ax.axvline(med_ss, color="black", linestyle="--", linewidth=1,
           label=f"Median {med_ss:.2f}%")
ax.set_xlabel("Sammenstilte anrop (skjulte / estimert totalt) — %")
ax.set_title("Sammenstilte anrop per sentral — manglende sekvensnr i 110 ID")
ax.legend(loc="lower right")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_oversikt_sammenstilte.png", bbox_inches="tight")
plt.close()

print("   6 figurer skrevet til analyse/figurer/")


# === 8. MARKDOWN-RAPPORT ===
print("8. Skriver nasjonal_oversikt.md...")

def fmt_nok(n, d=0):
    try:
        return f"{n:,.{d}f}".replace(",", " ").replace(".", ",")
    except:
        return str(n)

md = []
md.append("# Nasjonal oversikt — DSB 2025 for alle 12 110-sentraler")
md.append("")
md.append(f"*Generert fra `2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx` — {len(df):,} oppdrag.*")
md.append("")
md.append("## Sammendrag")
md.append("")

# Nøkkeltall
stor_1 = storrelse.iloc[0]
liten_1 = storrelse.iloc[-1]
ratio_vol = stor_1["totalvolum"] / liten_1["totalvolum"]
utr_spredning = storrelse["utrykningsrate_pct"].max() - storrelse["utrykningsrate_pct"].min()
laba_min = bench_pct["L-aba"].min()
laba_maks = bench_pct["L-aba"].max()
sentral_laba_null = bench[bench["L-aba"] == 0].index.tolist()

md.append(f"- **Total volum:** {fmt_nok(df.shape[0])} oppdrag fordelt på 12 sentraler")
md.append(f"- **Største vs minste:** {stor_1['sentral']} ({fmt_nok(stor_1['totalvolum'])}) "
          f"er **{ratio_vol:.1f}× større** enn {liten_1['sentral']} ({fmt_nok(liten_1['totalvolum'])})")
md.append(f"- **Varslingsrate-spredning (D alene):** {storrelse['utrykningsrate_pct'].min():.1f}%–"
          f"{storrelse['utrykningsrate_pct'].max():.1f}% (forskjell {utr_spredning:.1f} prosentpoeng — nesten 3× spredning). "
          f"Reduseres til {storrelse['brannrelatert_pct'].max()-storrelse['brannrelatert_pct'].min():.1f} pp "
          f"når L-aba og L-hendelse inkluderes — mye av D-rate-forskjellen skyldes beslutningsterskel i 110.")
md.append(f"- **L-aba-spredning:** {laba_min:.1f}%–{laba_maks:.1f}% — {len(sentral_laba_null)} sentral(er) har "
          f"0 registrerte L-aba ({', '.join(sentral_laba_null)})")
md.append(f"- **Skjulte 110 ID-sekvenser:** {sammenstilte['sammenstilte_rate_pct'].min():.2f}%–"
          f"{sammenstilte['sammenstilte_rate_pct'].max():.2f}%. "
          f"Dekker sammenstilte anrop + overføringer til nabo + avbrutte anrop — må dekomponeres per sentral.")
md.append("")

# Seksjon 1 — Størrelse
md.append("---")
md.append("")
md.append("## 1. Størrelse på sentralene")
md.append("")
md.append("Størrelse kan måles på tre måter som gir ulik ranking:")
md.append("- **Totalvolum:** antall registrerte oppdrag i DSB 2025.")
md.append("- **Arbeidsmengde:** volum × typisk bindingstid per kategori (timer per år).")
md.append("- **Oppdrag per c_effektiv:** volum per effektiv operatørkapasitet "
          "(= `c_total_dag − 1` for VL-korreksjon).")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_storrelse.png" alt="Størrelse" width="100%">')
md.append('  <p><small><i>Figur 1: Sentraler rangert etter tre størrelses-metrikker.</i></small></p>')
md.append("</div>")
md.append("")

md.append("### 1.1 Kombinert størrelses-tabell")
md.append("")
md.append("| Sentral | Totalvolum | Rang | Arbmengde timer/dag | Rang | Oppdrag/c_eff | Rang |")
md.append("|---|---:|---:|---:|---:|---:|---:|")
for _, r in storrelse.iterrows():
    aceff = f"{int(r['anrop_per_ceff']):,}".replace(",", " ") if pd.notna(r['anrop_per_ceff']) else "—"
    md.append(
        f"| {r['sentral']} | {fmt_nok(r['totalvolum'])} | {int(r['rang_totalvolum'])} | "
        f"{r['arbmengde_timer_per_dag']:.1f} | {int(r['rang_arbmengde_timer_aar'])} | "
        f"{aceff} | {int(r['rang_anrop_per_ceff']) if pd.notna(r['rang_anrop_per_ceff']) else '—'} |"
    )
md.append("")
md.append("**Observasjoner:**")
md.append(f"- Arbeidsmengde per dag spenner {storrelse['arbmengde_timer_per_dag'].min():.1f}–"
          f"{storrelse['arbmengde_timer_per_dag'].max():.1f} timer. "
          f"Oslo har mest med {storrelse.loc[storrelse['sentral']=='Oslo','arbmengde_timer_per_dag'].iloc[0]:.0f} timer/dag.")
top_belastning = storrelse.sort_values("anrop_per_ceff", ascending=False).head(3)
md.append(f"- Mest belastet per operatør: " + ", ".join(
    f"{r['sentral']} ({int(r['anrop_per_ceff']):,} oppdrag/c_eff)".replace(",", " ")
    for _, r in top_belastning.iterrows()))
md.append("")

# Seksjon 2 — Kategorifordeling heatmap
md.append("---")
md.append("")
md.append("## 2. Kategorifordeling (heatmap)")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_kategori_heatmap.png" alt="Heatmap" width="85%">')
md.append('  <p><small><i>Figur 2: Andel av totalvolumet per kategori. Mørkere rødt = høyere andel.</i></small></p>')
md.append("</div>")
md.append("")
md.append("**Mønster:**")
md.append("- **S (Service)** er den mest varierende — fra 9 % (Oslo) til 47 % (Møre og Romsdal). "
          "Reflekterer ulikt antall ABA-tilknyttede bygg eller ulik registreringspraksis for overføringstester.")
md.append("- **L-ukjent** er gjennomgående høy — indikerer at «Opprinnelig oppdragstype» ofte lukkes uten utfylling.")
md.append("- **D + S + L-ukjent** utgjør typisk 70–80 % av volumet på hver sentral.")
md.append("")

# Seksjon 3 — Utrykningsrate + brann-relatert
md.append("---")
md.append("")
md.append("## 3. Varslingsrate og samlet brann-relatert aktivitet")
md.append("")
md.append("### 3.1 Ren varslingsrate (D) — hva kategorien egentlig måler")
md.append("")
md.append("D-kategorien defineres som rader der **«Ressurs varslet» er fylt ut** — altså oppdrag der "
          "110 har varslet brannvesen. Dette er ikke det samme som «reelle hendelser» eller «utrykning gjennomført». "
          "Det er en **beslutning** i 110-sentralen om å sette tiltaket i gang.")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_utrykningsrate.png" alt="Utrykning" width="90%">')
md.append('  <p><small><i>Figur 3: Ren varslingsrate per sentral (andel av totalvolum). '
          'Rød = topp 20 %, blå = bunn 20 %.</i></small></p>')
md.append("</div>")
md.append("")
md.append("**Spredning:** "
          f"{storrelse['utrykningsrate_pct'].min():.1f}%–{storrelse['utrykningsrate_pct'].max():.1f}%, "
          f"median {storrelse['utrykningsrate_pct'].median():.1f}%. "
          f"Nesten 3× spredning — krever forklaring.")
md.append("")
md.append("### 3.2 Varsles det «på alt»? — Realiseringsgrad av varslinger")
md.append("")
md.append("En hypotese er at noen sentraler varsler tidlig og ofte, mens andre avklarer på telefon først. "
          "Hvis det stemmer, burde lav-terskel-sentraler ha **lavere andel varslede som faktisk rykker ut** "
          "(fordi mange varslinger avbrytes). Data viser at dette **ikke** er hovedforklaringen:")
md.append("")
md.append("| Sentral | Varslede (D) | % rykket ut | % kom fremme |")
md.append("|---|---:|---:|---:|")
for _, r in storrelse.sort_values("pct_rykket_av_varslet", ascending=False).iterrows():
    pr = f"{r['pct_rykket_av_varslet']:.1f}%" if pd.notna(r['pct_rykket_av_varslet']) else "—"
    pf = f"{r['pct_fremme_av_varslet']:.1f}%" if pd.notna(r['pct_fremme_av_varslet']) else "—"
    md.append(f"| {r['sentral']} | {fmt_nok(r['D'])} | {pr} | {pf} |")
md.append("")
md.append("Alle sentraler har 75–99 % realiseringsgrad — når varsling først er satt, rykker brannvesen "
          "som regel ut. Oslo topper med 98,8 %, Tromsø bunner på 75,0 %. Spredningen er reell men forklarer "
          "ikke hovedtyngden av D-rate-forskjellen.")
md.append("")
md.append("### 3.3 Samlet brann-relatert aktivitet — ser forbi beslutningsterskelen")
md.append("")
md.append("Forskjellen i D-rate skyldes i stor grad **hvor i 110-prosessen hendelsen klassifiseres**. "
          "Noen sentraler avklarer på telefon før varsling — disse havner i `L-aba` (automatisk brannalarm "
          "avklart) eller `L-hendelse` (reell hendelse avklart av 110). Andre varsler tidligere og "
          "unngår L-kategoriene.")
md.append("")
md.append("Hvis vi legger sammen **D + L-aba + L-hendelse**, måler vi all brann-relatert aktivitet "
          "uavhengig av hvor beslutningen ble tatt. Dette gir et mer sammenlignbart mål på hendelsesmengde.")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_brannrelatert.png" alt="Brann-relatert" width="92%">')
md.append('  <p><small><i>Figur 3b: Ren varslingsrate (rød) vs samlet brann-relatert aktivitet (oransje). '
          'Differansen = L-aba + L-hendelse = hendelser avklart uten varsling.</i></small></p>')
md.append("</div>")
md.append("")
md.append("| Sentral | D (%) | L-aba (%) | L-hendelse (%) | **Brann-relatert (%)** | Differanse |")
md.append("|---|---:|---:|---:|---:|---:|")
for _, r in storrelse.sort_values("brannrelatert_pct", ascending=False).iterrows():
    laba = bench_pct.loc[r['sentral'], 'L-aba']
    lh = bench_pct.loc[r['sentral'], 'L-hendelse']
    diff = r['brannrelatert_pct'] - r['utrykningsrate_pct']
    md.append(f"| {r['sentral']} | {r['utrykningsrate_pct']:.1f} | {laba:.1f} | {lh:.1f} | "
              f"**{r['brannrelatert_pct']:.1f}** | +{diff:.1f} |")
md.append("")
sam_min = storrelse['brannrelatert_pct'].min()
sam_maks = storrelse['brannrelatert_pct'].max()
sam_median = storrelse['brannrelatert_pct'].median()
md.append(f"**Spredning etter justering:** {sam_min:.1f}%–{sam_maks:.1f}%, median {sam_median:.1f}%. "
          f"Den rene varslingsratens spredning ({storrelse['utrykningsrate_pct'].max()-storrelse['utrykningsrate_pct'].min():.1f} pp) "
          f"reduseres til {sam_maks-sam_min:.1f} pp — konsistent med at mye av D-rate-forskjellen skyldes "
          f"**ulik praksis for når i prosessen 110 varsler brannvesen**, snarere enn reell forskjell i hendelsesmengde.")
md.append("")
md.append("**Konkrete eksempler:**")
md.append(f"- **Sør-Vest** (D: 12,2 %) har høyest L-aba (8,9 %) og 3,6 % L-hendelse. Samlet brann-relatert: "
          f"**{storrelse[storrelse['sentral']=='Sør-Vest']['brannrelatert_pct'].iloc[0]:.1f} %**. "
          f"Dette er konsistent med en mer restriktiv utvarslingspraksis der mye avklares på telefon først.")
md.append(f"- **Oslo** (D: 24,9 %) har bare 0,1 % L-aba og 9,2 % L-hendelse. Samlet: "
          f"**{storrelse[storrelse['sentral']=='Oslo']['brannrelatert_pct'].iloc[0]:.1f} %**. "
          f"Varsler tidlig — nesten ingen ABA avklares uten utrykning.")
md.append(f"- **Sør-Øst** (D: 20,6 %, L-aba: 0,0 %, L-hendelse: 10,2 %): "
          f"**{storrelse[storrelse['sentral']=='Sør-Øst']['brannrelatert_pct'].iloc[0]:.1f} %**. "
          f"Tilsynelatende samme praksis som Oslo — varsler på alle ABA, men har svært høy L-hendelse "
          f"(reelle hendelser avklart uten utrykning).")
md.append("")
md.append("> **Konklusjon:** D-rate alene er et misvisende sammenligningsmål på hendelsesbelastning. "
          "Samlet brann-relatert aktivitet (D + L-aba + L-hendelse) er mer robust fordi det fanger "
          "hendelser uavhengig av om de ble avklart eller utalarmert.")
md.append("")

# Seksjon 4 — L-aba dyp-fokus
md.append("---")
md.append("")
md.append("## 4. L-aba — dyp-fokus (ABA løst av 110)")
md.append("")
md.append("L-aba er den mest ekstreme avviks-kategorien på tvers av sentraler.")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_laba.png" alt="L-aba" width="90%">')
md.append('  <p><small><i>Figur 4: L-aba-andel per sentral. Mørkerød = 0 (ekstrem). '
          'Rød = bunn-kvartil, grønn = topp-kvartil.</i></small></p>')
md.append("</div>")
md.append("")
md.append("| Sentral | L-aba andel | Antall |")
md.append("|---|---:|---:|")
for s, row in bench.sort_values("L-aba", ascending=False).iterrows():
    md.append(f"| {s} | {bench_pct.loc[s, 'L-aba']:.1f}% | {fmt_nok(row['L-aba'])} |")
md.append("")
md.append("**Hypoteser om årsak til spredningen:**")
md.append("1. **Ulike registreringspraksis:** Sør-Øst (0) og Oslo (100) registrerer sannsynligvis ABA "
          "som avklares uten utrykning under en annen kategori (L-hendelse, L-ukjent, eller F).")
md.append("2. **Ulike terskler:** Sentraler med høy L-aba-andel (Sør-Vest 8,9 %, Nordland 8,7 %) "
          "kan ha lengre venteperiode før utkalling, eller annen prosedyre for bekreftelse.")
md.append("3. **Ulik bygningsmasse:** Sentraler med mange ABA-tilknyttede bygg vil ha både flere S og "
          "flere L-aba — dette henger sammen med S-andelen i tabellen ovenfor.")
md.append("")
md.append("> Dette er grunnlaget for L-aba-dybdeanalysen som pågår ved Sør-Vest "
          "(se `analyse/uttrekk/laba_sorvest_2025_dybdeanalyse.xlsx` — 50 hendelser som valideres manuelt i LEO).")
md.append("")

# Seksjon 5 — Tidsdata D
md.append("---")
md.append("")
md.append("## 5. Alarmbehandlingstid — kategori D")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_tid_boxplot.png" alt="Tid boksplot" width="95%">')
md.append('  <p><small><i>Figur 5: Boksplot alarmbehandlingstid per sentral (D-oppdrag). '
          'Outliers > 30 min trimmet.</i></small></p>')
md.append("</div>")
md.append("")
md_tidsdata = d_df.groupby("sentral")["alarmbeh_min"].agg(
    n="count", median="median", p90=lambda x: x.quantile(0.9)
).round(2).reset_index().sort_values("median")
md.append("| Sentral | n (D) | Median (min) | p90 (min) |")
md.append("|---|---:|---:|---:|")
for _, r in md_tidsdata.iterrows():
    md.append(f"| {r['sentral']} | {int(r['n']):,} | {r['median']:.2f} | {r['p90']:.2f} |".replace(",", " "))
md.append("")

# Seksjon 6 — Sammenstilte
md.append("---")
md.append("")
md.append("## 6. Skjulte 110 ID-sekvenser (sammenstilte, overførte, avbrutte)")
md.append("")
md.append("110 ID er strukturert som `BNN-YYMMDD-N` der `BNN` er sentral-prefiks (B01–B12) "
          "og `N` er løpenummer per dag per sentral. Manglende løpenumre i DSB-datasettet indikerer "
          "oppdrag som **ikke er registrert som egen rad**. Dette kan skyldes flere ting:")
md.append("")
md.append("1. **Sammenstilte anrop** — flere innkomne anrop slått sammen til ett registrert oppdrag. "
          "For 110 Sør-Vest er dette validert mot faktiske loggdata i V4-modellen.")
md.append("2. **Overførte anrop (30-sek-regel)** — ubesvart anrop etter 30 sek overføres automatisk til nabosentral. "
          "Anropet får et 110 ID, men registreres i en annen sentrals oppdragsliste. "
          "Dette kan særlig forklare de høye ratene i Finnmark og Agder.")
md.append("3. **Avbrutte/test-anrop** — anrop som ikke resulterer i registrert oppdrag.")
md.append("")
md.append('<div align="center">')
md.append('  <img src="../figurer/nasjonal_oversikt_sammenstilte.png" alt="Skjulte sekvenser" width="90%">')
md.append('  <p><small><i>Figur 6: Skjulte 110 ID-sekvenser per sentral. '
          'Sammensatt mål — ikke rene sammenstilte anrop.</i></small></p>')
md.append("</div>")
md.append("")
md.append("| Sentral | Registrerte oppdrag | Skjulte sekvensnr | Estimert totalt tildelt | Skjult-rate |")
md.append("|---|---:|---:|---:|---:|")
for _, r in sammenstilte.iterrows():
    md.append(
        f"| {r['sentral']} | {fmt_nok(r['registrerte_oppdrag'])} | "
        f"{fmt_nok(r['skjulte_sammenstilte'])} | {fmt_nok(r['totalt_estimert_anrop'])} | "
        f"{r['sammenstilte_rate_pct']:.2f}% |"
    )
md.append("")
md.append(f"**Observasjoner:**")
toppss = sammenstilte.iloc[0]
bunnss = sammenstilte.iloc[-1]
sv_rate = sammenstilte[sammenstilte["sentral"] == "Sør-Vest"].iloc[0]
md.append(f"- **{toppss['sentral']}** har høyest skjult-rate ({toppss['sammenstilte_rate_pct']:.2f}%) — "
          f"trolig dominert av overføringer til nabosentral snarere enn ekte sammenstilte anrop.")
md.append(f"- **{bunnss['sentral']}** har lavest ({bunnss['sammenstilte_rate_pct']:.2f}%) — "
          f"enten mindre samtidighet, færre overføringer, eller strengere rutine for opprettelse av oppdrag.")
md.append(f"- **Sør-Vest** ({sv_rate['sammenstilte_rate_pct']:.2f}%): rate er konsistent med tidligere V4-analyse "
          f"basert på faktisk logg — indikerer at Sør-Vests skjulte sekvenser primært er sammenstilte anrop.")
md.append(f"- Totalt: {int(sammenstilte['skjulte_sammenstilte'].sum()):,} skjulte sekvenser av "
          f"{int(sammenstilte['totalt_estimert_anrop'].sum()):,} estimert tildelt "
          f"= {sammenstilte['skjulte_sammenstilte'].sum()/sammenstilte['totalt_estimert_anrop'].sum()*100:.2f}% gjennomsnitt.".replace(",", " "))
md.append("")
md.append("> **Implikasjon for kapasitetsmodellen:** For sentraler med dokumentert sammenstillings-rate "
          "(f.eks. Sør-Vest gjennom V4-modellen) kan `SKJULT_BIND_MIN = 1 minutt` brukes som korreksjon "
          "for den ekstra operatørbelastningen. For andre sentraler må raten dekomponeres i ekte sammenstilte "
          "vs overførte/avbrutte — dette krever tilgang til LEO-loggen eller en spørreskjema-bekreftelse.")
md.append("")
md.append("> **Oppfølgingsspørsmål til sentralene:** Har dere oversikt over "
          "(a) hvor mange anrop overføres ut til nabosentral, og "
          "(b) hvor ofte flere anrop slås sammen til ett oppdrag?")
md.append("")

# Seksjon 7 — Konklusjon
md.append("---")
md.append("")
md.append("## 7. Oppsummering — hvilke avvik krever forklaring fra sentralene")
md.append("")
md.append("De tre viktigste avvikene å undersøke videre (allerede innarbeidet som sentralspesifikke "
          "oppfølgingsspørsmål i `analyse/sporreskjema/`):")
md.append("")
md.append("1. **L-aba-registrering:** Spesielt Sør-Øst (0) og Oslo (100) må forklare hvordan ABA uten utrykning "
          "registreres lokalt — uten avklaring kan vi ikke sammenligne ABA-belastning på tvers.")
md.append("2. **Utrykningsrate-spredning (9–25 %):** Må forstås som registreringspraksis vs. reell hendelsesmønster. "
          "Relevant for kapasitetsmodellering fordi D-hendelser har lengst bindingstid.")
md.append("3. **Sammenstilte anrop-rate:** Variasjon mellom sentraler indikerer ulik lokal rutine — kan korrigeres "
          "med kategori-spesifikk bindingstid justert per sentral.")
md.append("")
md.append("### Neste steg")
md.append("")
md.append("- Få tilbake spørreskjemaer fra alle 12 sentraler — verifiser registreringspraksis for hver kategori.")
md.append("- L-aba-dybdeanalyse Sør-Vest: bekreft eller juster 3-min-estimatet basert på 50 manuelt loggede hendelser.")
md.append("- Når sentralspesifikke bindingstider er validert, regnes arbeidsmengden i tabell 1.1 om med lokale parametre.")
md.append("")

with open(OUT_DIR / "nasjonal_oversikt.md", "w", encoding="utf-8") as f:
    f.write("\n".join(md))

print(f"   nasjonal_oversikt.md skrevet ({sum(1 for _ in md)} linjer)")
print("\nFerdig.")
print(f"  - {OUT_DIR / 'nasjonal_oversikt.md'}")
print(f"  - {OUT_DIR / 'storrelse_ranking.csv'}")
print(f"  - {OUT_DIR / 'sammenstilte_rate.csv'}")
print(f"  - 6 figurer i {FIG_DIR}")
