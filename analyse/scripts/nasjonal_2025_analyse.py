"""
Nasjonal analyse av DSB-fullrapport 2025 — alle 12 norske 110-sentraler.

Input: 004 data/2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx (508k rader)

Output:
  - analyse/nasjonal_2025/benchmarkmatrise.csv
  - analyse/nasjonal_2025/volumavstemming.csv
  - analyse/nasjonal_2025/kategori_per_sentral.csv (andeler)
  - analyse/nasjonal_2025/tidsdata_kategori_D.csv
  - analyse/nasjonal_2025/anomalier.md
  - analyse/figurer/nasjonal_2025_kategori_andeler.png
  - analyse/figurer/nasjonal_2025_utrykningsandel.png
  - analyse/figurer/nasjonal_2025_tid_per_sentral.png

Kjoer: py analyse/scripts/nasjonal_2025_analyse.py
"""
import pathlib
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import seaborn as sns

warnings.filterwarnings("ignore")

# === KONFIGURASJON ===
PROJECT = pathlib.Path(__file__).resolve().parent.parent.parent
DATA_DIR = PROJECT / "004 data"
OUT_DIR = PROJECT / "analyse" / "nasjonal_2025"
FIG_DIR = PROJECT / "analyse" / "figurer"
OUT_DIR.mkdir(parents=True, exist_ok=True)
FIG_DIR.mkdir(parents=True, exist_ok=True)

DSB_FIL = DATA_DIR / "2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx"
TIDL_FULL = DATA_DIR / "20260315_174350_fullrapport.csv"
MOB_2025 = DATA_DIR / "20260315_174514_MOB_2025_110-sentral.xlsx"
SORVEST_FASIT = PROJECT / "analyse" / "total_belastning_kategorisering.csv"

sns.set_style("whitegrid")
plt.rcParams.update({"figure.dpi": 150, "font.size": 10})


# === NAVN-NORMALISERING ===
SENTRAL_NORM = {
    "Sør-Vest 110": "Sør-Vest",
    "S\u00f8r-Vest 110": "Sør-Vest",
    "S?r-Vest 110": "Sør-Vest",
    "Oslo 110": "Oslo",
    "Øst 110": "Øst",
    "\u00d8st 110": "Øst",
    "?st 110": "Øst",
    "Sør-Øst 110": "Sør-Øst",
    "S\u00f8r-\u00d8st 110": "Sør-Øst",
    "S?r-?st 110": "Sør-Øst",
    "Vest 110": "Vest",
    "Midt-Norge 110": "Midt-Norge",
    "Innlandet 110": "Innlandet",
    "Agder 110": "Agder",
    "Nordland 110": "Nordland",
    "Møre og Romsdal 110": "Møre og Romsdal",
    "M\u00f8re og Romsdal 110": "Møre og Romsdal",
    "M?re og Romsdal 110": "Møre og Romsdal",
    "Tromsø 110": "Tromsø",
    "Troms\u00f8 110": "Tromsø",
    "Troms? 110": "Tromsø",
    "Finnmark 110": "Finnmark",
}
ALLE_SENTRALER = sorted(set(SENTRAL_NORM.values()))


def norm_sentral(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s in SENTRAL_NORM:
        return SENTRAL_NORM[s]
    for k, kanon in SENTRAL_NORM.items():
        if s.lower().startswith(kanon.lower()):
            return kanon
    return s


# === 1. LAST DSB-FIL ===
print("=" * 70)
print("1. LAST DSB-FIL")
print("=" * 70)

# Prøv ulike skiprows for å finne ekte header
for sr in [0, 1, 2, 3]:
    try:
        test = pd.read_excel(DSB_FIL, engine="openpyxl", skiprows=sr, nrows=3)
        cols = [str(c) for c in test.columns]
        if any("Oppdrag" in c for c in cols):
            print(f"Bruker skiprows={sr}")
            DSB_SKIPROWS = sr
            break
    except Exception as e:
        continue
else:
    DSB_SKIPROWS = 2

df = pd.read_excel(DSB_FIL, engine="openpyxl", skiprows=DSB_SKIPROWS)
print(f"Lastet: {len(df):,} rader, {len(df.columns)} kolonner")
print(f"Kolonner: {list(df.columns)[:8]}...")

# Normaliser kolonnenavn
col_map = {}
for c in df.columns:
    cs = str(c).strip()
    col_map[c] = cs
df = df.rename(columns=col_map)

# Finn nøkkelkolonner via fuzzy match
def finn_kol(kandidater, required=True):
    for cand in kandidater:
        for c in df.columns:
            if str(c).lower().replace(" ", "").replace("_", "") == cand.lower().replace(" ", "").replace("_", ""):
                return c
    for cand in kandidater:
        for c in df.columns:
            if cand.lower().replace(" ", "").replace("_", "") in str(c).lower().replace(" ", "").replace("_", ""):
                return c
    if required:
        raise KeyError(f"Fant ikke kolonne for {kandidater}. Tilgjengelig: {list(df.columns)}")
    return None

K_SENTRAL = finn_kol(["110-sentral", "110sentral"])
K_OPPDRAGTYPE = finn_kol(["Oppdragstype"])
K_OPPRINNELIG = finn_kol(["Opprinnelig oppdragstype", "Opprinneligoppdragstype"])
K_OVERORDNET = finn_kol(["Overordnet oppdragstype"], required=False)
K_RESSURS_VARSLET = finn_kol(["Ressurs varslet", "Ressursvarslet"], required=False)
K_OPPDRAG_ID = finn_kol(["Oppdrag ID", "OppdragID"], required=False)
K_110_ID = finn_kol(["110 ID", "110ID"], required=False)
K_ALARMBEH = finn_kol(["Alarmbehandlingstid"], required=False)
K_UTRYKNINGSTID = finn_kol(["Utrykningstid"], required=False)
K_RESPONSTID = finn_kol(["Responstid"], required=False)
K_KILDE = finn_kol(["Kilde"], required=False)

print(f"Sentral-kolonne:         {K_SENTRAL}")
print(f"Oppdragstype:            {K_OPPDRAGTYPE}")
print(f"Opprinnelig oppdragstype: {K_OPPRINNELIG}")
print(f"Ressurs varslet:         {K_RESSURS_VARSLET}")
print(f"Alarmbehandlingstid:     {K_ALARMBEH}")

df["sentral"] = df[K_SENTRAL].apply(norm_sentral)
print(f"\nUnike sentraler: {sorted(df['sentral'].dropna().unique())}")

# Konverter tidskolonner
if K_RESSURS_VARSLET:
    df[K_RESSURS_VARSLET] = pd.to_datetime(df[K_RESSURS_VARSLET], errors="coerce")


# === 2. V3-KATEGORISERING ===
print("\n" + "=" * 70)
print("2. V3-KATEGORISERING")
print("=" * 70)

def klassifiser(row):
    """V3-kat: D/S/L-aba/L-hendelse/L-ukjent/F/V
    Oppdatert 2026-04-19: L-aba krever Kilde=Alarm. ABA-oppdrag med
    Kilde=Samtale eller blank reklassifiseres til L-hendelse."""
    if K_RESSURS_VARSLET and pd.notna(row[K_RESSURS_VARSLET]):
        return "D"

    ot = str(row[K_OPPDRAGTYPE]).strip() if pd.notna(row[K_OPPDRAGTYPE]) else ""
    oot = str(row[K_OPPRINNELIG]).strip() if pd.notna(row[K_OPPRINNELIG]) else ""
    kilde = ""
    if K_KILDE and pd.notna(row[K_KILDE]):
        kilde = str(row[K_KILDE]).strip()

    if ot == "Service":
        return "S"

    feilring = {"Nødanrop feilring", "Ikke reell nødmelding",
                "ECall feil bruk", "ECall teknisk/ukjent", "ECall veihjelp",
                "Nodanrop feilring", "Ikke reell nodmelding"}
    if ot in feilring or "feilring" in ot.lower() or "ikke reell" in ot.lower():
        return "F"

    if "viderevarslet" in ot.lower() or "viderekoble" in ot.lower():
        return "V"

    if "ppdrag" in ot and "110" in ot:
        if oot == "ABA" and kilde == "Alarm":
            return "L-aba"
        elif oot == "ABA":
            return "L-hendelse"
        elif oot and oot.lower() != "nan":
            return "L-hendelse"
        else:
            return "L-ukjent"

    if ot == "" and oot == "":
        return "L-ukjent"

    return "L-ukjent"

df["v3_kat"] = df.apply(klassifiser, axis=1)

kat_totalt = df["v3_kat"].value_counts()
print("Nasjonalt totalt:")
for k in ["D", "S", "L-aba", "L-hendelse", "L-ukjent", "F", "V"]:
    n = kat_totalt.get(k, 0)
    print(f"  {k:12s}: {n:>8,} ({n/len(df)*100:>5.1f}%)")
print(f"  {'TOTAL':12s}: {len(df):>8,}")


# === 3. BENCHMARKMATRISE PER SENTRAL ===
print("\n" + "=" * 70)
print("3. BENCHMARKMATRISE PER SENTRAL")
print("=" * 70)

bench = df.groupby(["sentral", "v3_kat"]).size().unstack(fill_value=0)
for k in ["D", "S", "L-aba", "L-hendelse", "L-ukjent", "F", "V"]:
    if k not in bench.columns:
        bench[k] = 0
bench = bench[["D", "S", "L-aba", "L-hendelse", "L-ukjent", "F", "V"]]
bench["Totalt"] = bench.sum(axis=1)

# Andeler
bench_pct = bench.div(bench["Totalt"], axis=0).drop(columns="Totalt") * 100
bench_pct.columns = [f"{c}_pct" for c in bench_pct.columns]

bench_full = pd.concat([bench, bench_pct], axis=1).reset_index()
bench_full.to_csv(OUT_DIR / "benchmarkmatrise.csv", index=False, encoding="utf-8-sig")
print(f"Skrevet: benchmarkmatrise.csv")
print(bench[["D", "L-aba", "L-hendelse", "L-ukjent", "F", "V", "Totalt"]].to_string())


# === 4. VOLUMAVSTEMMING ===
print("\n" + "=" * 70)
print("4. VOLUMAVSTEMMING (DSB vs tidligere fullrapport vs MOB)")
print("=" * 70)

# Tidligere fullrapport
try:
    tidl = pd.read_csv(TIDL_FULL, sep=None, engine="python", skiprows=2,
                       encoding="utf-8-sig")
    hdr = tidl.iloc[0].tolist()
    tidl.columns = hdr
    tidl = tidl.iloc[1:].reset_index(drop=True)
    tidl["sentral"] = tidl["110-sentral"].apply(norm_sentral)
    tidl_vol = tidl.groupby("sentral").size().to_dict()
    print(f"Tidligere fullrapport lastet: {len(tidl):,} rader")
except Exception as e:
    print(f"FEIL ved lasting av tidligere fullrapport: {e}")
    tidl_vol = {}

# MOB 2025
try:
    import openpyxl
    wb = openpyxl.load_workbook(MOB_2025, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r[0] and "110-sentral" in str(r[0]))
    hdrs = [str(h).strip() if h else "" for h in rows[hdr_idx]]
    anrop_idx = next(i for i, h in enumerate(hdrs) if "mottatte" in h.lower() and "110-anrop" in h.lower())
    mob_vol = {}
    for r in rows[hdr_idx+1:]:
        if not r[0] or "Sum" in str(r[0]):
            continue
        s = norm_sentral(r[0])
        if r[anrop_idx] is not None:
            mob_vol[s] = int(float(r[anrop_idx]))
    print(f"MOB 2025 lastet: {len(mob_vol)} sentraler")
except Exception as e:
    print(f"FEIL ved MOB-lasting: {e}")
    mob_vol = {}

dsb_vol = df.groupby("sentral").size().to_dict()

rows = []
for s in ALLE_SENTRALER:
    dsb = dsb_vol.get(s, 0)
    tidl_n = tidl_vol.get(s, 0)
    mob = mob_vol.get(s, 0)
    dsb_vs_tidl = ((dsb - tidl_n) / tidl_n * 100) if tidl_n else np.nan
    dsb_vs_mob = ((dsb - mob) / mob * 100) if mob else np.nan
    rows.append({
        "sentral": s,
        "DSB_ny": dsb,
        "Tidl_fullrapport": tidl_n,
        "MOB_selvrapport": mob,
        "DSB_vs_tidl_pct": round(dsb_vs_tidl, 1) if not pd.isna(dsb_vs_tidl) else None,
        "DSB_vs_MOB_pct": round(dsb_vs_mob, 1) if not pd.isna(dsb_vs_mob) else None,
    })
volavst = pd.DataFrame(rows)
volavst.to_csv(OUT_DIR / "volumavstemming.csv", index=False, encoding="utf-8-sig")
print(volavst.to_string(index=False))


# === 5. KONSISTENSSJEKK SØR-VEST MOT FASIT ===
print("\n" + "=" * 70)
print("5. KONSISTENSSJEKK SØR-VEST")
print("=" * 70)

try:
    fasit = pd.read_csv(SORVEST_FASIT, encoding="utf-8-sig")
    print(f"Fasit (gammel kategorisering) lastet: {len(fasit):,} rader")
    print(f"Fasit kategorifordeling:")
    print(fasit["v3_kat"].value_counts())

    sv_ny = df[df["sentral"] == "Sør-Vest"]
    print(f"\nNy kategorisering Sør-Vest: {len(sv_ny):,} rader")
    print("Ny kategorifordeling:")
    print(sv_ny["v3_kat"].value_counts())

    # Match på Oppdrag ID (hvis samme IDer finnes)
    if K_OPPDRAG_ID and "Oppdrag_ID" in fasit.columns:
        ny_ids = sv_ny[K_OPPDRAG_ID].astype(str).str.strip()
        fasit_ids = fasit["Oppdrag_ID"].astype(str).str.strip()
        overlapp = set(ny_ids) & set(fasit_ids)
        print(f"\nFelles Oppdrag IDer: {len(overlapp):,}")
        if overlapp:
            # Slå sammen og sammenlign v3_kat
            sv_merge = sv_ny[[K_OPPDRAG_ID, "v3_kat"]].copy()
            sv_merge[K_OPPDRAG_ID] = sv_merge[K_OPPDRAG_ID].astype(str).str.strip()
            sv_merge = sv_merge.rename(columns={"v3_kat": "v3_ny", K_OPPDRAG_ID: "Oppdrag_ID"})
            fasit_merge = fasit[["Oppdrag_ID", "v3_kat"]].copy()
            fasit_merge["Oppdrag_ID"] = fasit_merge["Oppdrag_ID"].astype(str).str.strip()
            fasit_merge = fasit_merge.rename(columns={"v3_kat": "v3_fasit"})
            comp = sv_merge.merge(fasit_merge, on="Oppdrag_ID", how="inner")
            match = (comp["v3_ny"] == comp["v3_fasit"]).sum()
            print(f"Match på ID: {match}/{len(comp)} ({match/len(comp)*100:.1f}%)")
            if match < len(comp):
                diff = comp[comp["v3_ny"] != comp["v3_fasit"]]
                print(f"Eksempler på avvik:")
                print(diff.head(10).to_string(index=False))
                print(f"\nForskjell-krysstabell:")
                print(pd.crosstab(diff["v3_fasit"], diff["v3_ny"]))
except Exception as e:
    print(f"Sjekk kunne ikke fullføres: {e}")


# === 6. TIDSDATA FOR KATEGORI D ===
print("\n" + "=" * 70)
print("6. TIDSDATA PER SENTRAL (kun kategori D)")
print("=" * 70)

d_df = df[df["v3_kat"] == "D"].copy()

def as_min(series):
    """Konverter tidskolonne (timedelta eller string) til minutter."""
    # timedelta64 (alle unit-varianter)
    if pd.api.types.is_timedelta64_dtype(series):
        return series.dt.total_seconds() / 60
    # String
    if series.dtype == object:
        parsed = pd.to_timedelta(series.astype(str), errors="coerce")
        return parsed.dt.total_seconds() / 60
    # Numeric fallback
    n = pd.to_numeric(series, errors="coerce")
    if n.max() and n.max() > 1000:
        return n / 60
    return n

tid_rows = []
for s in ALLE_SENTRALER:
    sub = d_df[d_df["sentral"] == s]
    row = {"sentral": s, "n_D": len(sub)}
    for kol, navn in [(K_ALARMBEH, "alarmbeh"), (K_UTRYKNINGSTID, "utrykn"), (K_RESPONSTID, "respons")]:
        if kol and kol in sub.columns:
            m = as_min(sub[kol])
            m = m.dropna()
            row[f"{navn}_n"] = len(m)
            row[f"{navn}_median_min"] = round(m.median(), 2) if len(m) else None
            row[f"{navn}_p90_min"] = round(m.quantile(0.9), 2) if len(m) else None
    tid_rows.append(row)
tidsdata = pd.DataFrame(tid_rows)
tidsdata.to_csv(OUT_DIR / "tidsdata_kategori_D.csv", index=False, encoding="utf-8-sig")
print(tidsdata.to_string(index=False))


# === 7. AVVIKSDETEKSJON ===
print("\n" + "=" * 70)
print("7. AVVIKSDETEKSJON (IQR-metode)")
print("=" * 70)

def finn_outliers(series, navn):
    q1, q3 = series.quantile([0.25, 0.75])
    iqr = q3 - q1
    lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    out = []
    for s, v in series.items():
        if v < lo or v > hi:
            side = "LAV" if v < lo else "HØY"
            out.append(f"- **{s}**: {navn} = {v:.1f}% ({side}; norm {lo:.1f}–{hi:.1f}%)")
    return out

bench_pct_nom = bench_pct.rename(columns=lambda c: c.replace("_pct", ""))
avvik_rader = []
for k in ["D", "L-aba", "L-ukjent", "F", "V"]:
    navn = {"D": "utrykningsrate", "L-aba": "ABA-andel",
            "L-ukjent": "L-ukjent-andel", "F": "feilring-andel",
            "V": "viderekoble-andel"}[k]
    outs = finn_outliers(bench_pct_nom[k], navn)
    avvik_rader.extend(outs)


# === 8. FIGURER ===
print("\n" + "=" * 70)
print("8. FIGURER")
print("=" * 70)

# Figur 1: stablet søyle kategoriandeler per sentral
sentraler_sortert = bench_pct_nom.sort_values("D", ascending=False).index.tolist()
fig, ax = plt.subplots(figsize=(13, 7))
kat_farger = {"D": "#C44E52", "S": "#8172B2", "L-aba": "#DD8452",
              "L-hendelse": "#55A868", "L-ukjent": "#CCB974",
              "F": "#64B5CD", "V": "#937860"}
bottom = np.zeros(len(sentraler_sortert))
for k in ["D", "L-aba", "L-hendelse", "L-ukjent", "S", "F", "V"]:
    v = bench_pct_nom.loc[sentraler_sortert, k].values
    ax.bar(sentraler_sortert, v, bottom=bottom, label=k, color=kat_farger[k])
    bottom += v
ax.set_ylabel("Prosent av totalvolum")
ax.set_title("Nasjonal fordeling av V3-kategorier per 110-sentral (DSB 2025)")
ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
plt.xticks(rotation=35, ha="right")
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_2025_kategori_andeler.png")
plt.close()
print("Skrevet: nasjonal_2025_kategori_andeler.png")

# Figur 2: utrykningsandel sortert
fig, ax = plt.subplots(figsize=(11, 6))
utr = bench_pct_nom["D"].sort_values(ascending=True)
farger = ["#C44E52" if v > bench_pct_nom["D"].quantile(0.75) + 1.5*(bench_pct_nom["D"].quantile(0.75)-bench_pct_nom["D"].quantile(0.25))
          else "#4C72B0" for v in utr.values]
ax.barh(utr.index, utr.values, color=farger)
ax.set_xlabel("Utrykningsrate (D / totalt, %)")
ax.set_title("Utrykningsrate per 110-sentral (DSB 2025)")
ax.axvline(utr.median(), color="grey", linestyle="--", label=f"Median {utr.median():.1f}%")
ax.legend()
plt.tight_layout()
plt.savefig(FIG_DIR / "nasjonal_2025_utrykningsandel.png")
plt.close()
print("Skrevet: nasjonal_2025_utrykningsandel.png")

# Figur 3: Median alarmbeh-tid per sentral (D)
if "alarmbeh_median_min" in tidsdata.columns:
    fig, ax = plt.subplots(figsize=(11, 6))
    t = tidsdata.set_index("sentral")["alarmbeh_median_min"].dropna().sort_values()
    ax.barh(t.index, t.values, color="#4C72B0")
    ax.set_xlabel("Median alarmbehandlingstid — kategori D (minutter)")
    ax.set_title("Alarmbehandlingstid (D) per 110-sentral (DSB 2025)")
    plt.tight_layout()
    plt.savefig(FIG_DIR / "nasjonal_2025_tid_per_sentral.png")
    plt.close()
    print("Skrevet: nasjonal_2025_tid_per_sentral.png")


# === 9. ANOMALI-RAPPORT ===
print("\n" + "=" * 70)
print("9. ANOMALI-RAPPORT")
print("=" * 70)

mdlines = []
mdlines.append("# Nasjonal analyse DSB 2025 — avvik og mønster")
mdlines.append("")
mdlines.append(f"**Datasett:** `2025_fullrapport_110_alle_sentraler_fra_dsb.xlsx`")
mdlines.append(f"**Totalvolum:** {len(df):,} oppdrag / {len(ALLE_SENTRALER)} sentraler")
mdlines.append("")
mdlines.append("## 1. Volumavstemming")
mdlines.append("")
mdlines.append("| Sentral | DSB ny | Tidl. fullrapport | MOB selvrapport | Δ DSB/tidl % | Δ DSB/MOB % |")
mdlines.append("|---|---:|---:|---:|---:|---:|")
for _, r in volavst.iterrows():
    mdlines.append(
        f"| {r['sentral']} | {r['DSB_ny']:,} | {r['Tidl_fullrapport']:,} | {r['MOB_selvrapport']:,} | "
        f"{r['DSB_vs_tidl_pct']} | {r['DSB_vs_MOB_pct']} |"
    )
mdlines.append("")
mdlines.append("**Observasjoner:**")
volavst["DSB_vs_tidl_pct_num"] = pd.to_numeric(volavst["DSB_vs_tidl_pct"], errors="coerce")
store_avvik = volavst[volavst["DSB_vs_tidl_pct_num"].abs() > 10]
if len(store_avvik):
    mdlines.append(f"- {len(store_avvik)} sentraler har > 10 % avvik mellom ny DSB-fil og tidligere fullrapport:")
    for _, r in store_avvik.iterrows():
        mdlines.append(f"  - {r['sentral']}: DSB {r['DSB_ny']:,} vs tidl. {r['Tidl_fullrapport']:,} ({r['DSB_vs_tidl_pct_num']:+.1f}%)")
else:
    mdlines.append("- Alle sentraler har < 10 % avvik mellom DSB og tidligere fullrapport (eller tidligere fullrapport mangler).")
mob_mean = volavst.apply(lambda r: r["DSB_ny"]/r["MOB_selvrapport"] if r["MOB_selvrapport"] else None, axis=1).dropna().mean()
mdlines.append("")
mdlines.append(f"- Gjennomsnittlig DSB-volum er {mob_mean:.1f}× MOB-selvrapportert anropsvolum.")
mdlines.append("  Dette bekrefter at DSB inkluderer kategorier MOB ikke teller som 'anrop' (løst av 110, viderekoblinger, feilringer).")
mdlines.append("")

mdlines.append("## 2. Kategorifordeling per sentral")
mdlines.append("")
mdlines.append("| Sentral | D % | L-aba % | L-hendelse % | L-ukjent % | F % | V % | Totalt |")
mdlines.append("|---|---:|---:|---:|---:|---:|---:|---:|")
for s in sentraler_sortert:
    r = bench_pct_nom.loc[s]
    total = int(bench.loc[s, "Totalt"])
    mdlines.append(
        f"| {s} | {r['D']:.1f} | {r['L-aba']:.1f} | {r['L-hendelse']:.1f} | "
        f"{r['L-ukjent']:.1f} | {r['F']:.1f} | {r['V']:.1f} | {total:,} |"
    )
mdlines.append("")

mdlines.append("## 3. Avvik (utenfor ±1,5 × IQR)")
mdlines.append("")
if avvik_rader:
    mdlines.extend(avvik_rader)
else:
    mdlines.append("- Ingen sentraler ligger utenfor IQR-terskelen på de vurderte forholdene.")
mdlines.append("")

mdlines.append("## 4. Tidsdata for kategori D (utrykning)")
mdlines.append("")
mdlines.append("| Sentral | n (D) | Median alarmbeh (min) | p90 alarmbeh | Median utrykn | Median respons |")
mdlines.append("|---|---:|---:|---:|---:|---:|")
for _, r in tidsdata.iterrows():
    ab_m = r.get("alarmbeh_median_min", "")
    ab_p = r.get("alarmbeh_p90_min", "")
    ut_m = r.get("utrykn_median_min", "")
    rs_m = r.get("respons_median_min", "")
    mdlines.append(f"| {r['sentral']} | {r['n_D']:,} | {ab_m} | {ab_p} | {ut_m} | {rs_m} |")
mdlines.append("")

mdlines.append("## 5. Hypoteser og oppfølgingsspørsmål")
mdlines.append("")
mdlines.append("Følgende spørsmål bør drøftes med sentralene eller legges til spørreskjemaene:")
mdlines.append("")
mdlines.append("- **Stor L-ukjent-andel:** Hvilken praksis ligger bak å lukke oppdrag uten `Opprinnelig oppdragstype`? Er dette nærmere korte samtaler (F-lignende) eller reelle hendelser som ikke kategoriseres?")
mdlines.append("- **ABA-variasjon:** Hva driver forskjellen i ABA-andel mellom sentraler? Ulike alarmleverandører, lokal prosedyre for utrykning på automatiske brannalarmer, eller ulik bygningsmasse?")
mdlines.append("- **Utrykningsrate:** Sentraler med høy utrykningsrate kan ha færre 'ufarlige' oppdrag som filtreres ut — eller mindre samlet volum slik at utrykninger får større relativ vekt. Sjekk mot innbyggertall og areal.")
mdlines.append("- **Alarmbehandlingstid:** Store forskjeller mellom sentraler kan indikere ulik registreringspraksis, ulike LEO-rutiner, eller reelle kapasitetsforskjeller. Må valideres mot sentralenes egne tall.")
mdlines.append("")

(OUT_DIR / "anomalier.md").write_text("\n".join(mdlines), encoding="utf-8")
print(f"Skrevet: anomalier.md ({len(mdlines)} linjer)")

print("\n" + "=" * 70)
print("FERDIG")
print("=" * 70)
print(f"Output i: {OUT_DIR}")
